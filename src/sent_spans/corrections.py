import logging
import operator
from pathlib import Path
import pickle

import openpyxl.reader.excel
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Alignment
import pandas as pd
from tqdm import tqdm

from src import make_classification_datasets

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
here = Path(__file__).parent

"""
Work-in-progress way of improving dataset quality.

Prepares a spreadsheet showing our sentence span case classification datasets, along with initial predictions 
from a BERT classification model. The idea is we can apply manual corrections and then re-import.
"""

UNCORRECTED_PATH = here / f'../data/results/bert_uncased_v2.1/case_predictions_211222.xlsx'
CORRECTED_VERSION = 'corrected_5-28-23'
CORRECTED_PATH = UNCORRECTED_PATH.parent / f'case_predictions_211222_{CORRECTED_VERSION}.xlsx'

SENT_SPAN_ORIGIN = here / f'../data/sent_span_classification_211222_v2.pkl'
DOC_ORIGIN =here / f'../data/doc_classification_211222_v2.pkl'
SENT_SPAN_OUT = SENT_SPAN_ORIGIN.parent / (str(SENT_SPAN_ORIGIN.stem) + f'_{CORRECTED_VERSION}.pkl')
DOC_OUT = DOC_ORIGIN.parent / (str(DOC_ORIGIN.stem) + f'_{CORRECTED_VERSION}.pkl')

SOURCES = [
    ('approved', 'positive', 'Approved points that the model is at least a little skeptical about -- potential false positives'),
    ('declined', 'negative', 'Declined points -- potential true negatives'),
    ('topical', 'negative', 'Approved points of *other* cases in the same topic area -- potential true negatives'),
    ('surrounding', 'negative', 'Surrounding sentences of approved points -- potential true negatives'),
    ('doc_random', 'negative', 'Random sentences from docs of approved points -- potential true negatives'),
    ('reviewed_random', 'negative', 'Random sentences from comprehensively reviewed docs -- potential true negatives')
]
GAP_LINES = 3

def make_xlsx(datasets, cases, threshold=0.03):
    wb = Workbook()

    for sheet_i, case_id in tqdm(enumerate(sorted(datasets.keys())), total=len(datasets)):
    # for sheet_i, case_id in [(0, 118)]:
        logger.info(f"Creating case {case_id}")
        worksheet = wb.create_sheet() if sheet_i > 0 else wb.active
        worksheet.title = str(case_id)
        worksheet.append(['Title:', cases.loc[case_id].title])
        worksheet.append(['Description:', cases.loc[case_id].description])
        worksheet.append([''])
        worksheet.append(['We\'re training a model to predict this case, but it\'s making mistakes. Sometimes it\'s just a  '
                          'tricky edge case, other times it\'s a mistakenly labeled training example. The purpose of this spreadsheet '
                          'is to correct the latter by double-checking contentious cases.'])
        worksheet.append([''])
        worksheet.append(['Instructions:'])
        worksheet.append(['- Write your name on row 12 to claim this case.'])
        worksheet.append(['- Modify our_label as needed to reflect the true status of the point. A "positive" label means '
                          'the text contains sufficient evidence to support the case (extra surrounding text is fine).'])
        worksheet.append(['- Points are grouped by method/reason used to include them in the dataset. They are sorted '
                          'by how contentious they are, started with ones that the model disagrees with the strongest.'])
        worksheet.append(['- You don\'t have to go through every row; you can give up beyond a certain ml_score (some very '
                          'confident answers are even hidden from this spreadsheet altogether)'])
        worksheet.append([''])
        worksheet.append(['Curator assignee:'])
        worksheet.append(['Thoughts:'])

        for source, _, desc in SOURCES:
            for __ in range(GAP_LINES):
                worksheet.append([''])
            cell = Cell(column='A', row=worksheet.max_row, value=f"type: {desc}", worksheet=worksheet)
            cell.font = Font(bold=True)
            worksheet.append([cell])
            worksheet.append(['type', 'document_id', 'char_start', 'char_end', 'ml_score', 'our_label', 'text', 'link'])

            instances = datasets[case_id][datasets[case_id].source == source]
            ascending = source == 'approved'
            for idx, row in instances.sort_values('pred', ascending=ascending).iterrows():
                pred = float(row.pred)
                if (source == 'approved' and pred >= (1. - threshold)) or \
                        (source != 'approved' and pred <= threshold):
                    logger.debug(f"Skipping {row.char_end} {source} {pred}")
                    continue
                else:
                    logger.debug(f"Cool with {row.char_end} {source} {pred}")
                if 'point_id' in row and not pd.isna(row.point_id):
                    link = f'https://edit.tosdr.org/points/{int(row.point_id)}'
                else:
                    link = f'https://edit.tosdr.org/services/{int(row.service_id)}/annotate'
                link_cell = Cell(worksheet=worksheet, value=f'=HYPERLINK("{link}", "{link}")', column=7, row=worksheet.max_row)
                cells = [
                    row.source, row.document_id, row.char_start, row.char_end, round(pred, 4), row.label, row.text, link_cell
                ]
                worksheet.append(cells)

        # Set text wrapping on the text (column G)
        for row in worksheet[2:worksheet.max_row]:
            row[6].alignment = Alignment(wrap_text=True)

    wb.save(UNCORRECTED_PATH)

def read_xlsx(path=CORRECTED_PATH):
    wb = openpyxl.reader.excel.load_workbook(path, read_only=True)

    source_names = set(map(operator.itemgetter(0), SOURCES))
    case_corrections = dict()
    for case_id in wb.sheetnames:
        corrections: set[tuple[int, int, int]] = set()

        source_idx = -1
        for idx, row in enumerate(wb[case_id].iter_rows(max_col=6, values_only=True)):
            if row[0] is not None:
                if row[0].startswith('type: '):
                    source_idx += 1
                if source_idx >= 0 and row[0] in source_names:
                    assert row[5] in {'positive', 'negative'}
                    if row[5] != SOURCES[source_idx][1]:
                        corrections.add(tuple(row[:4]))

        for source in sorted(source_names):
            num_corr = len(list(filter(lambda corr: corr[0] == source, corrections)))
            if num_corr > 0:
                print(f"Case {case_id}: {num_corr} {source} corrections")
        case_corrections[int(case_id)] = corrections
    return case_corrections

def apply_corrections(
        correction_path,
        sent_span_input_path, sent_span_output_path,
        doc_input_path, doc_output_path,
        # To test the effect of corrections, we can only apply them to fold 0 which will be used as evaluation
        # (at least for sent spans; the doc dataset is eval only and can be fully corrected)
        restrict_to_fold=None
):
    # We want to be very careful not to overwrite data, since make_classification_datasets.py is nondeterministic
    # (at least the 211222 dataset can't be repro'd)
    assert sent_span_input_path != sent_span_output_path
    assert doc_input_path != doc_output_path

    orig_labels = dict([(source_info[0], source_info[1]) for source_info in SOURCES])
    _flip_label = lambda label: 'positive' if label == 'negative' else 'negative'

    sent_span_df: dict[pd.DataFrame] = pickle.load(open(sent_span_input_path, 'rb'))
    corrections = read_xlsx(correction_path)
    for case_id in corrections:
        df = sent_span_df[case_id]
        for correction_tuple in corrections[case_id]:
            source, doc_id, start_char, end_char = correction_tuple
            # Not efficient at all, but that's ok
            matches = df[(df.source == source) & (df.document_id == doc_id) & (df.char_start == start_char)]
            if len(matches) == 0:
                raise ValueError(f"Correction {correction_tuple} not found for case ID {case_id}")
            if len(matches) > 1:
                logger.warning(f"Multiple origins found for correction {correction_tuple} case ID {case_id}")
            for idx, match in matches.iterrows():
                if restrict_to_fold is None or match.fold == restrict_to_fold:
                    sent_span_df[case_id].at[idx, 'label'] = _flip_label(orig_labels[source])

    logger.info(f"Wrote corrected sent spans to {sent_span_output_path}")
    pickle.dump(sent_span_df, open(sent_span_output_path, 'wb'))

    # Also correct doc datasets. This won't be perfect because there might be negative examples we haven't corrected,
    # but some of the sent span corrections will make it in so it's better than nothing.
    doc_df: dict[pd.DataFrame] = pickle.load(open(doc_input_path, 'rb'))
    for case_id in corrections:
        num_changed = 0
        df = doc_df[case_id]
        for source, doc_id, start_char, end_char in corrections[case_id]:
            # Approved points that were changed to negative could still have positive spans elsewhere in the doc,
            # but negative points that were changed to positive means a potential change in doc label.
            if orig_labels[source] == 'negative':
                matches = df[df.id_doc == doc_id]
                if len(matches) > 0:
                    if len(matches) > 1:
                        raise ValueError(f"Multiple doc id {doc_id} in doc dataset for {case_id}?")
                    if doc_df[case_id].at[matches.iloc[0].name, 'label'] == 'negative':
                        num_changed += 1
                    doc_df[case_id].at[matches.iloc[0].name, 'label'] = 'positive'
        if num_changed > 0:
            logger.info(f"Doc dataset case {case_id}: applied {num_changed} corrections")

    logger.info(f"Wrote corrected docs to {doc_output_path}")
    pickle.dump(doc_df, open(doc_output_path, 'wb'))


if __name__ == '__main__':
    # datasets = pickle.load(open(RESULTS_DIR / 'sent_span_pred.pkl', 'rb'))
    # cases = pickle.load(open(here / f'../data/cases_{VERSION}_clean.pkl', 'rb'))
    # make_xlsx(datasets, cases)

    apply_corrections(CORRECTED_PATH, SENT_SPAN_ORIGIN, SENT_SPAN_OUT, DOC_ORIGIN, DOC_OUT)
    # apply_corrections(
    #     CORRECTED_PATH,
    #     SENT_SPAN_ORIGIN,
    #     SENT_SPAN_ORIGIN.parent / (str(SENT_SPAN_ORIGIN.stem) + f'_{CORRECTED_VERSION}_evalonly.pkl'),
    #     DOC_ORIGIN,
    #     DOC_ORIGIN.parent / (str(DOC_ORIGIN.stem) + f'_{CORRECTED_VERSION}_evalonly.pkl'),
    #     restrict_to_fold=0
    # )
